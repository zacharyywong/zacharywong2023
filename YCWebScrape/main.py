# This is a sample Python script.

# Press ⌃R to execute it or replace it with your code.
# Press Double ⇧ to search everywhere for classes, files, tool windows, actions, and settings.
from datetime import datetime
import pytz as pytz
from timeit import *


pst = pytz.timezone('US/Pacific')
fmt = '%d-%m-%Y'
now = datetime.now()
now = now.astimezone(pst).strftime(fmt)

# Import utility packages
import concurrent.futures
import logging
import queue
import threading

from openpyxl import Workbook
import pandas as pd
import numpy as np
import time
from time import sleep
import pytz
from openpyxl.utils.dataframe import dataframe_to_rows

# import selenium/webscraping libs
from selenium import webdriver
#from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC

from WebScrape import *
from Excel import *
# Global Variables

# Access webdriver for Chrome
driver_path = './chromedriver'
service = Service(driver_path)
driver = None
waitTime = 15
mainURL = 'https://www.ycombinator.com/companies?batch=S22&regions=America%20%2F%20Canada'
scrape = WebScrape(waitTime, 6, 8)
linksList = []
namesList = []
drivers = []
failedCompanyList = []

companyLinkDF = pd.DataFrame()
companyLinkSheetName = 'Links'



companyInfoColumns = ["Name", "Website", "YC Link", "Short Description", "Long Description", "Categories", "Founded Year",
                      "Team Size", "Location"]
companyInfoWSName = 'Company Info'
companyInfoDF = pd.DataFrame(columns = companyInfoColumns)

companyFounderColumns = ['Company Name', 'Founder Name', 'LinkedIn', 'Description (If Available)']
companyFounderDF = pd.DataFrame(columns = companyFounderColumns)
companyFounderWSName = 'Founder Info'
companyInfoTotalFileName = 'YC Demo Company Info.xlsx'
excel = Excel(companyInfoTotalFileName)
wb = None

mutex = threading.Lock()


def loadCreateWorkbook():
    global wb, companyLinkDF, driver
    try:
        wb = excel.loadWorkbook(companyInfoTotalFileName)
        ws = wb[companyLinkSheetName]
        companyLinkDF = excel.excelToDF(ws)
        #print(companyLinkDF)
        return True
    except FileNotFoundError:
        print('Cannot load workbook - does not exist')
        print('creating a new workbook')
        wb = excel.createWorkbook()
        driver = webdriver.Chrome(service=Service(driver_path))
        driver.get(mainURL)
        return False

def writetoExcel(df, sheetName, fileName):
    excel.writetoExcel(df, sheetName, fileName, index=False, header=True)

def getCompanyNumber():
    global driver
    numberXPATH = '/html/body/div/div[2]/section[2]/div/div[2]/div[3]'
    driver, companyNumberElement = scrape.getElementXPATH(driver, numberXPATH)
    companyNumber = companyNumberElement.text.split(' ')[3]
    print("Number of Companies: " + str(int(companyNumber)))
    return int(companyNumber)

def getAllCompanyURL():
    global linksList, companyLinkDF, driver, namesList

    companyNumber = getCompanyNumber()

    XPATH = '//*[@class="styles-module__section___2yul1 styles-module__results___2lP37"]'

    #TODO:
    #number of companies
    #numberofCompanies = scrape.getElementXPATH(driver, '//*[@id="CompaniesIndexPage-react-component-055f0f5a-16b8-4fcf-81d1-baef61aeec89"]/div[2]/section[2]/div/div[1]/div/div[9]/div[3]/label/span[2]')
    #scroll to bottom of page

    driver, companyAll = scrape.getElementXPATH(driver, XPATH)

    while len(linksList) < companyNumber:
        companyWEList = companyAll.find_elements(By.XPATH, "//*[@class='styles-module__company___1UVnl no-hovercard']")
        for link in [company.get_attribute('href') for company in companyWEList if company.get_attribute('href') not in linksList]:
            linksList.append(str(link))
       # nameElements = companyAll.find_elements(By.XPATH, '//*[@class="styles-module__coName___3zz21"]')
       # for name in [name.text for name in nameElements if name.text not in namesList]:
       #     namesList.append(name)

        driver.execute_script('window.scrollBy(0,document.body.scrollHeight)')
    #print(len(namesList))
   # print(len(linksList))
    print(namesList)
    print(linksList)
    #companyLinkDF = pd.DataFrame(columns = ['Name', 'YC Demo Links'])
    companyLinkDF = pd.DataFrame(columns = ['YC Demo Links'])

    #companyLinkDF['Name'] = namesList
    companyLinkDF['YC Demo Links'] = linksList
    driver.quit()

def getWebsiteElement(driver, XPATH):
    driver, element = scrape.getElementXPATH(driver, XPATH)
    element = element.text
    return element

def getFounderInformation(driver, link):
    global expandedForm
    #global driver

    # Founders names
    try:
        founderParentXPATH = '//div[@class="space-y-4"]'
        driver, founderParentElement = scrape.getElementXPATH(driver, founderParentXPATH)
        expandedForm = False


    except BaseException:
        #print("Founders Section in Expanded Form")
        founderParentXPATH = '//div[@class="space-y-5"]'
        driver, founderParentElement = scrape.getElementXPATH(driver, founderParentXPATH)
        expandedForm = True
    try:
        founderNames = []
        founderElements = founderParentElement.find_elements(By.XPATH, '//div[@class="ycdc-card shrink-0 space-y-1.5 sm:w-[300px]"]')
        for founderE in founderElements:
            founderNames.append(founderE.text.split('\n', 1)[0])
        #print(founderNames)
    except:
        print("Cannot Find Founder Names for %s.. returning empty founder information" % link)
        return None

    try:
        # founder LinkedIns
        founderLinkedIns = []
        linkedInXPATH = '//div[@class="mt-1 space-x-2"]/a'
        founderElements = (founderParentElement.find_elements(By.XPATH, linkedInXPATH))
        [founderLinkedIns.append(e.get_attribute('href')) for e in founderElements if 'linkedin' in e.get_attribute('href')]
        #print(founderLinkedIns)
        # get descriptions of each founder
        founderDescriptions = []
        if expandedForm:
            descriptionXPATH = '//div[@class="flex-grow"]/p'
            founderDescElements = founderParentElement.find_elements(By.XPATH, descriptionXPATH)
            [founderDescriptions.append(e.text) for e in founderDescElements]
            #print(founderDescriptions)
        else:
            [founderDescriptions.append("") for index in range(len(founderNames))]
            #print("Founder Descriptions: " + str(founderDescriptions))
        # make each founder a list within list
        founderInfos = []
        for index in range(len(founderNames)):
            founderInfo = []
            founderInfo.append(founderLinkedIns[index])
            founderInfo.append(founderDescriptions[index])
            founderInfos.append(founderInfo)
        #print(founderInfos)
        founderInfoDict = dict(zip(founderNames, founderInfos))
        return founderInfoDict
            #founder -> linkedin
        #print(founderLinkedIns)
        # create founder linkedIn Dictionary
    except:
        print("Cannot find LinkedIn Information... returning dict with keys but no value")
        return founderInfoDict

def fillCompanyInfoRow(tempDF, name, shortDesc, longDesc, websiteURL, categoryList, yearFounded, teamSize, location):
    companyInfoColumns = ["Name", "Website", "Short Description", "Long Description", "Categories", "Founded Year", "Team Size", "Location"]
    tempDF['Name'] = name
    tempDF['Website'] = websiteURL
    tempDF['Short Description'] = shortDesc
    tempDF['Long Description'] = longDesc
    tempDF['Categories'] = categoryList
    tempDF['Founded Year'] = yearFounded
    tempDF['Team Size'] = teamSize
    tempDF['Location'] = location
    return tempDF


def getCompanyInformation(index):
    global companyInfoDF, companyFounderDF, linksList, mutex, failedCompanyList
    #global driver
    #print("working on %s" % link)
    # lock thread when popping from linksList
    with mutex:
        logging.info("locking thread %d" % index)
        try:
            link = linksList.pop(0)
            driver = drivers[index]
        except:
            logging.info("all links and/or drivers already retrieved.. returning")
            return
        logging.info("retrieved and removed link from linksList: %s" % link)
        logging.info("retrieved driver from driversList: %s" % index)
        print("linksList: " + str(linksList))
        logging.info("unlocking thread %d" % index)
    try:
        driver.get(link)
    except BaseException:
        #failedCompanyList.append(link)
        print("cannot get YC page for %s" % link)
        return

    name = None
    shortDesc = None
    longDesc = None
    websiteURL= None
    categoryList= None
    yearFounded= None
    teamSize= None
    location= None

    try:
        #Company Name
        nameXPATH = '//div[@class="prose max-w-full"]/h1'
        name = getWebsiteElement(driver, nameXPATH)
        #print("Name: " + name)
    except:
        print("Cannot Find Company Name for %s" % link)

    #Short Description
    try:
        shortDescXPATH = '//div[@class="text-xl"]'
        shortDesc = getWebsiteElement(driver, shortDescXPATH)
        #print('Short Desc: ' + shortDesc)
    except:
        print("Cannot Find Short Description for %s" % link)

    # Long Description
    try:
        longDescXPATH = '//p[@class="whitespace-pre-line"]'
        longDesc = getWebsiteElement(driver, longDescXPATH)
        #print('Long Desc: ' + longDesc)
    except:
        print("Cannot Find Long Description for %s" % link)

    # Company Website
    try:
        websiteURLXPATH = '//div[@class="flex flex-row items-center leading-none px-3"]/a'
        driver, websiteURLElement = scrape.getElementXPATH(driver, websiteURLXPATH)
        websiteURL = websiteURLElement.get_attribute('href')
        #print('URL: ' + websiteURL)
    except:
        print("Cannot Find Company Website for %s" % link)

    try:
        currentURL = driver.current_url
    except:
        print("Cannot get current URL for %s" % link)

    # Company Categories
    try:
        categoryList = []
        XPATHCategories = '/html/body/div/div[2]/section/div[2]/div[1]/div[1]/div[2]/div[3]'
        driver, categoryParent = scrape.getElementXPATH(driver, XPATHCategories)
        childElementXPATH = "//span[@class = 'ycdc-badge']"
        try:
            categories = categoryParent.find_elements(By.XPATH, childElementXPATH)
        except:
            WebDriverWait(driver, waitTime).until(EC.presence_of_element_located((By.XPATH, childElementXPATH)))
            categories = categoryParent.find_elements(By.XPATH, childElementXPATH)

        [categoryList.append(category.text) for category in categories]
        #print("Category List: " + str(categoryList))

    except:
        print("Cannot Find Company Categories for %s" % link)

    # Founded Year, Team Size, Location
    try:
        result = []
        parentXPATH = '/html/body/div[1]/div[2]/section/div[2]/div[2]/div[1]/div[2]'
        driver, parent = scrape.getElementXPATH(driver, parentXPATH)
        list = parent.find_elements(By.XPATH, "//div[@class = 'flex flex-row justify-between']")
        for text in [e.text.split('\n', 1)[1] for e in list]:
            result.append(text)
        yearFounded = result[-3]
        teamSize = result[-2]
        location = result[-1]
        #print("Year Founded is %s / Team Size is %s / Location is %s" % (yearFounded, teamSize, location))
    except:
        print("Cannot Find Year, Team Size, and/or Location for %s" % link)

    founderInfoDict = getFounderInformation(driver, link)

    # lock thread when updating company/founder dataframe
    with mutex:
        logging.info("locking thread %d" % index)
        logging.info("updating company info to df for %s" % link)
        companyInfoDF.loc[len(companyInfoDF.index)] = [name, websiteURL, currentURL, shortDesc, longDesc, str(categoryList),
                                                       yearFounded, teamSize, location]
        logging.info("updating founder names to df for %s" % link)
        for founder in founderInfoDict.keys():
            companyFounderDF.loc[len(companyFounderDF.index)] = [name, founder, founderInfoDict.get(founder)[0], founderInfoDict.get(founder)[1]]
        logging.info("unlocking thread %d" % index)

    print("done with %s" % link)

# run all helper functions to scrape info from Google
def runThreads(linksList, numThreads):
    global drivers
    # list of dictionaries with key/value pairs: title, link, text
    # Contains all information for all search results
    ##(numThreads)
    #with concurrent.futures.ThreadPoolExecutor(max_workers=numThreads) as executor:
    #   # chunksize = math.ceil(len(linksList) / numThreads)
    #    #print("chunksize: " + str(chunksize))
    #    for link in linksList:
    #       executor.submit(getCompanyInformation, link)
    #       executor.shutdown()

    # change links list to links queue
    #linkQueue = queue.Queue()
    #[linkQueue.put(link) for link in linksList]

    # create and start threads while running linksList is not finished
    threads = []
    for index in range(numThreads):
        driver = webdriver.Chrome(service=service)
        drivers.append(driver)
        time.sleep(1)
    while len(linksList) > 0:
        logging.info("Starting new thread and driver pool with thread number: %d" % numThreads)
        # create and append drivers

        for index in range(0, numThreads):
            # create and start threads
            logging.info("creating and starting thread: %d",  index)
            thread = threading.Thread(target = getCompanyInformation , args = (index, ))
            threads.append(thread)
            thread.start()

        print("starting this round of threads: " + str(threads))
        for index, thread in enumerate(threads):
            logging.info("Main thread: waiting for thread %d to finish", index)
            thread.join()
            logging.info("Main thread calling again: thread %d is finished", index)

        threads.clear()
        print("finished with this round of threads: " + str(threads))
    # quit and remove drivers
    logging.info("quitting drivers")
    [driver.quit() for driver in drivers]
    drivers.clear()

def checkFailedList():
    global failedCompanyList
    originalLinkList = companyLinkDF['YC Demo Links'].tolist()
    resultLinkList = companyInfoDF['YC Link'].tolist()
    print("originalLinkList: " + str(originalLinkList))
    print("resultLinkList: " + str(resultLinkList))

    [failedCompanyList.append(link) for link in originalLinkList if link not in resultLinkList]
    print(failedCompanyList)

def run():
    startTime = datetime.now()
    global linksList
    loadWorkbook = loadCreateWorkbook()
    if not loadWorkbook:
        getAllCompanyURL()
        writetoExcel(companyLinkDF, companyLinkSheetName, companyInfoTotalFileName)
        print("Number of Links to Scrape: " + str(len(linksList)))

    maxThreads = 5
    linksList = companyLinkDF['YC Demo Links'].tolist()
    print(linksList)
    numThreads = min(maxThreads, len(linksList))
    print('Number of Threads: ' + str(numThreads))
    logging.getLogger().setLevel(logging.INFO)
    runThreads(linksList, numThreads)
    print("company info df: " + str(companyInfoDF))
    print("company info df: " + str(companyFounderDF))
    excel.writetoExcel(companyInfoDF, companyInfoWSName, companyInfoTotalFileName, index=True, header=True)
    excel.writetoExcel(companyFounderDF, companyFounderWSName, companyInfoTotalFileName, index=True, header=True)

    checkFailedList()
    dfFailed = pd.DataFrame(failedCompanyList)
    failedWSName = 'Failed Scrape Companies'
    excel.writetoExcel(dfFailed, failedWSName, companyInfoTotalFileName, index = False, header = False)

    endTime = datetime.now()
    elapsedTime = endTime-startTime
    print("Total Time Elapsed: " + str(elapsedTime))
if __name__ == '__main__':
    run()