# Class for all functions related to writing / loading files from Excel using openpyxl library

from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

class Excel:
    wb = Workbook()

    def __init__(self, totalFileName):
        self.totalFileName = totalFileName

    def createWorkbook(self):
        global wb
        wb = Workbook()
        print("Created new workbook named %s" % (self.totalFileName))
        wb.save(self.totalFileName)
        return wb

    def loadWorkbook(self, totalFileName):
        global wb
        try:
            wb = load_workbook(totalFileName)
            print("loading %s workbook" % (totalFileName))
        except ImportError as e:
            print("check Excel file: most likely corrupt")
        return wb

    # Helper function to write a sheet to excel
    # :param DataFrame df: dataframe to save into excel
    # :param str sheetName: name of excel sheet to save to
    def writetoExcel(self, df, sheetName, totalFileName, index, header):
        ws = wb.create_sheet()
        try:
            del wb[sheetName]
            print("deleted " + sheetName)
        except LookupError:
            print("No duplicate sheets to delete")
        finally:
            ws.title = sheetName
            rows = dataframe_to_rows(df, index=index, header=header)
            print(rows)
            for row in rows:
                #if len(row) >= 1 and row[0] is not None:
                print(row)
                ws.append(row)
            wb.save(totalFileName)
            print("wrote {%s} to {%s}" % (sheetName, totalFileName))
            return wb

    def excelToDF(self, ws):
        data = ws.values
        columns = next(data)[0:]
        df = pd.DataFrame(data, columns=columns)
        return df
